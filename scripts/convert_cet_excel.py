"""Convert IPU CET Excel cutoff data to TypeScript."""
import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'd:\Download\IPU_CET_College_Predictor_2025.xlsx')

SHEET_COURSE_MAP = {
    '📘 BBA Data': 'BBA',
    '💻 BCA Data': 'BCA', 
    '⚖️ Law Data': 'Law',
    '💰 B.Com Data': 'B.Com',
    '🖥️ MCA Data': 'MCA',
    '🎓 MBA Data': 'MBA',
}

entries = []
for sheet_name, course_group in SHEET_COURSE_MAP.items():
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        vals = [cell.value for cell in row]
        sno, college, course, shift, category, open_rank, close_rank, remarks = vals[:8]
        if sno is None or not isinstance(sno, (int, float)):
            continue
        if not isinstance(open_rank, (int, float)) or not isinstance(close_rank, (int, float)):
            continue
        # Determine region
        cat_str = str(category or '').strip()
        if 'Outside Delhi' in cat_str:
            region = 'outside_delhi'
        elif 'Delhi' in cat_str:
            region = 'delhi'
        else:
            region = 'delhi'
        # Determine category type
        if 'SC' in cat_str:
            cat_type = 'SC'
        elif 'OBC' in cat_str:
            cat_type = 'OBC'
        else:
            cat_type = 'GEN'
        
        entries.append({
            'college': str(college or '').strip(),
            'course': str(course or '').strip(),
            'courseGroup': course_group,
            'shift': str(shift or '').strip(),
            'category': cat_type,
            'region': region,
            'openRank': int(open_rank),
            'closeRank': int(close_rank),
            'remarks': str(remarks or '').strip() if remarks else '',
        })

# Write TypeScript
lines = []
lines.append('// Auto-generated from IPU_CET_College_Predictor_2025.xlsx')
lines.append('// Last Round (Round 3) cutoffs for 2025-26')
lines.append('')
lines.append('export type CetCourseGroup = "BBA" | "BCA" | "Law" | "B.Com" | "MCA" | "MBA";')
lines.append('')
lines.append('export type CetCutoffEntry = {')
lines.append('  college: string;')
lines.append('  course: string;')
lines.append('  courseGroup: CetCourseGroup;')
lines.append('  shift: string;')
lines.append('  category: "GEN" | "SC" | "OBC";')
lines.append('  region: "delhi" | "outside_delhi";')
lines.append('  openRank: number;')
lines.append('  closeRank: number;')
lines.append('  remarks: string;')
lines.append('};')
lines.append('')
lines.append('export const cetCutoffData: CetCutoffEntry[] = ')
lines.append(json.dumps(entries, indent=2, ensure_ascii=False) + ';')

with open(r'd:\c drive\OneDrive\Desktop\Ipuhub\lib\cet-cutoff-data.ts', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines) + '\n')

print(f'Generated {len(entries)} entries')
for cg in SHEET_COURSE_MAP.values():
    count = sum(1 for e in entries if e['courseGroup'] == cg)
    print(f'  {cg}: {count}')
